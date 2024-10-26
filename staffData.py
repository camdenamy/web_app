import sqlite3
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import matplotlib.pyplot as plt
import re

# Define allowed names with categories
STAFF_CATEGORIES = {
    "Discord Support": {"Asher", "Ogea", "Jerome"},
    "Councilman": {"Tony", "Hanson", "Smith", "Knight", "Zen", "Riggs"},
    "Chairman": {"Amy", "Williams", "BMac"},
    "Commissioner": {"Fenix"},
    "Lieutenant Governor": {"Gibbs"},
    "Community Manager": {"Epik"},
    "Governor": {"Jimmy"}
}

CATEGORY_RANKS = [
    "Governor", "Community Manager", "Lieutenant Governor", "Commissioner", "Chairman", "Councilman"
]

# Ticket Class Definition
class Ticket:
    def __init__(self, ticket_number=None, date_of_ticket=None, ticket_type=None, answered_by=None,
                 response_time=None, claimed_by=None, closed_by=None, reviewed_by=None, handled=None, notes=None):
        self.ticket_number = ticket_number
        self.date_of_ticket = self.convert_to_datetime(date_of_ticket)
        self.ticket_type = ticket_type
        self.answered_by = answered_by
        self.response_time = self.convert_to_int(response_time)
        self.claimed_by = claimed_by
        self.closed_by = closed_by
        self.reviewed_by = reviewed_by
        self.handled = handled
        self.notes = notes

    @staticmethod
    def convert_to_datetime(date_input):
        if isinstance(date_input, str):
            try:
                return datetime.strptime(date_input, "%m/%d/%Y")
            except ValueError:
                return None
        return date_input

    @staticmethod
    def convert_to_int(time_input):
        try:
            return int(time_input)
        except (ValueError, TypeError):
            return 0

class ModeratorInteraction:
    def __init__(self, moderator_name, date_of_interaction, interaction_type):
        self.moderator_name = moderator_name
        self.date_of_interaction = self.convert_to_datetime(date_of_interaction)
        self.interaction_type = interaction_type

    @staticmethod
    def convert_to_datetime(date_input):
        if isinstance(date_input, str):
            try:
                return datetime.strptime(date_input, "%m/%d/%Y")
            except ValueError:
                return None
        return date_input

# Create or connect to the database
def create_database():
    conn = sqlite3.connect('staff_management.db')
    c = conn.cursor()

    # Create tickets and interactions table if they don't exist
    c.execute('''CREATE TABLE IF NOT EXISTS tickets (
                    ticket_number TEXT PRIMARY KEY,
                    date_of_ticket TEXT,
                    ticket_type TEXT,
                    answered_by TEXT,
                    response_time INTEGER,
                    claimed_by TEXT,
                    closed_by TEXT,
                    reviewed_by TEXT,
                    handled TEXT,
                    notes TEXT
                )''')

    c.execute('''CREATE TABLE IF NOT EXISTS interactions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    moderator_name TEXT,
                    date_of_interaction TEXT,
                    interaction_type TEXT
                )''')

    c.execute('''CREATE TABLE IF NOT EXISTS staff (
                    name TEXT PRIMARY KEY,
                    category TEXT
                )''')

    for category, names in STAFF_CATEGORIES.items():
        for name in names:
            c.execute("INSERT OR IGNORE INTO staff (name, category) VALUES (?, ?)", (name, category))

    conn.commit()
    conn.close()

# Insert ticket into the database
def insert_ticket(ticket):
    conn = sqlite3.connect('staff_management.db')
    c = conn.cursor()

    if ticket.date_of_ticket:
        ticket.date_of_ticket = Ticket.convert_to_mmddyyyy(ticket.date_of_ticket)

    c.execute('''INSERT INTO tickets (ticket_number, date_of_ticket, ticket_type, answered_by,
                                      response_time, claimed_by, closed_by, reviewed_by, handled, notes)
                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
              (ticket.ticket_number, ticket.date_of_ticket, ticket.ticket_type, ticket.answered_by,
               ticket.response_time, ticket.claimed_by, ticket.closed_by, ticket.reviewed_by,
               ticket.handled, ticket.notes))
    conn.commit()
    conn.close()

# Insert interaction into the database
def insert_interaction(interaction):
    conn = sqlite3.connect('staff_management.db')
    c = conn.cursor()

    if interaction.date_of_interaction:
        interaction.date_of_interaction = ModeratorInteraction.convert_to_mmddyyyy(interaction.date_of_interaction)

    c.execute('''INSERT INTO interactions (moderator_name, date_of_interaction, interaction_type)
                 VALUES (?, ?, ?)''',
              (interaction.moderator_name, interaction.date_of_interaction, interaction.interaction_type))
    conn.commit()
    conn.close()

# Check if ticket number already exists
def ticket_number_exists(ticket_number):
    conn = sqlite3.connect('staff_management.db')
    c = conn.cursor()
    c.execute("SELECT 1 FROM tickets WHERE ticket_number = ?", (ticket_number,))
    result = c.fetchone()
    conn.close()
    return result is not None

# Query to get all staff
def get_all_staff():
    conn = sqlite3.connect('staff_management.db')
    c = conn.cursor()
    c.execute("SELECT name, category FROM staff")
    staff = c.fetchall()
    conn.close()
    return staff

# Get filtered tickets based on moderator, month, and year
def get_filtered_tickets(moderator_name=None, selected_month=None, selected_year=None):
    conn = sqlite3.connect('staff_management.db')
    c = conn.cursor()
    query = '''SELECT * FROM tickets WHERE (answered_by = ? OR claimed_by = ?)'''
    params = [moderator_name, moderator_name]
    if selected_month and selected_year:
        query += " AND strftime('%m', date_of_ticket) = ? AND strftime('%Y', date_of_ticket) = ?"
        params.extend([selected_month, selected_year])
    c.execute(query, params)
    tickets = c.fetchall()
    conn.close()
    return tickets

# Query to get filtered interactions
def get_filtered_interactions(moderator_name, selected_month, selected_year):
    conn = sqlite3.connect('staff_management.db')
    c = conn.cursor()
    query = '''SELECT * FROM interactions WHERE moderator_name = ? AND 
               strftime('%m', date_of_interaction) = ? AND 
               strftime('%Y', date_of_interaction) = ?'''
    c.execute(query, (moderator_name, selected_month, selected_year))
    interactions = c.fetchall()
    conn.close()
    return interactions

# Function to calculate average response time
def calculate_average_response_time(tickets):
    total_response_time = 0
    ticket_count = 0

    # Iterate through tickets to calculate total response time
    for ticket in tickets:
        response_time = ticket[5]
        if response_time is not None and response_time != "":
            try:
                response_time = int(response_time)
                total_response_time += response_time
                ticket_count += 1
            except ValueError:
                print(f"Skipping invalid response time: {response_time}")

    # Calculate the average if there are valid tickets
    if ticket_count > 0:
        average_response_time = total_response_time / ticket_count
    else:
        average_response_time = 0

    return average_response_time

def upload_from_excel():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return

    try:
        workbook = openpyxl.load_workbook(file_path)

        # Process the 'Tickets' sheet if it exists
        if 'Tickets' in workbook.sheetnames:
            ticket_sheet = workbook['Tickets']
            for row in ticket_sheet.iter_rows(min_row=2, values_only=True):
                if row:
                    ticket_number = str(row[0]).strip() if row[0] else None

                    # Skip if ticket number is missing or already exists
                    if not ticket_number or ticket_number_exists(ticket_number):
                        continue

                    # Convert and normalize the date field
                    date_of_ticket = convert_excel_date(row[1])

                    ticket_type = normalize_field(row[2], "Unknown")
                    answered_by = normalize_field(row[3], "Unknown")
                    claimed_by = normalize_field(row[4], "Unknown")
                    response_time = row[5] if isinstance(row[5], int) else 0
                    closed_by = normalize_field(row[6], "Unknown")
                    reviewed_by = normalize_field(row[7], "Unknown")
                    handled = normalize_field(row[8], "No")
                    notes = normalize_field(row[9], "")

                    # Create and insert the ticket
                    ticket = Ticket(
                        ticket_number=ticket_number,
                        date_of_ticket=date_of_ticket,
                        ticket_type=ticket_type,
                        answered_by=answered_by,
                        claimed_by=claimed_by,
                        response_time=response_time,
                        closed_by=closed_by,
                        reviewed_by=reviewed_by,
                        handled=handled,
                        notes=notes
                    )
                    insert_ticket(ticket)

        # Process the 'Interactions' sheet if it exists
        if 'Interactions' in workbook.sheetnames:
            interaction_sheet = workbook['Interactions']
            for row in interaction_sheet.iter_rows(min_row=2, values_only=True):
                if row:
                    moderator_name = normalize_field(row[0], "Unknown")

                    # Convert and normalize the date field
                    date_of_interaction = convert_excel_date(row[1])

                    interaction_type = normalize_field(row[2], "General")

                    # Create and insert the interaction
                    interaction = ModeratorInteraction(
                        moderator_name=moderator_name,
                        date_of_interaction=date_of_interaction,
                        interaction_type=interaction_type
                    )
                    insert_interaction(interaction)

        messagebox.showinfo("Success", "Data uploaded successfully from Excel!")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to upload data: {str(e)}")

def normalize_field(value, default):
    return str(value).strip() if isinstance(value, str) else default

def convert_excel_date(date_value):
    """
    Converts various date formats from Excel to 'MM/DD/YYYY' format.
    """
    try:
        if isinstance(date_value, datetime):
            # If it's a datetime object, format it directly
            return date_value.strftime("%m/%d/%Y")
        elif isinstance(date_value, (int, float)):
            # If it's an Excel serial date, convert it to a date
            excel_date = datetime(1899, 12, 30) + timedelta(days=int(date_value))
            return excel_date.strftime("%m/%d/%Y")
        elif isinstance(date_value, str):
            # If it's a string, try to parse it in various formats
            try:
                return datetime.strptime(date_value, "%Y-%m-%d").strftime("%m/%d/%Y")
            except ValueError:
                pass
            try:
                return datetime.strptime(date_value, "%m/%d/%Y").strftime("%m/%d/%Y")
            except ValueError:
                pass
            try:
                return datetime.strptime(date_value, "%d/%m/%Y").strftime("%m/%d/%Y")
            except ValueError:
                return None
    except Exception as e:
        print(f"Error converting date: {date_value}, Error: {e}")
        return None

def insert_ticket(ticket):
    """
    Inserts a ticket into the database and prints the ticket details for verification.
    """
    conn = sqlite3.connect('staff_management.db')
    c = conn.cursor()

    # Debug: Print ticket details before insertion
    print(f"Inserting ticket: {ticket.ticket_number}, Date: {ticket.date_of_ticket}, Type: {ticket.ticket_type}")

    c.execute('''INSERT INTO tickets (ticket_number, date_of_ticket, ticket_type, answered_by,
                                      response_time, claimed_by, closed_by, reviewed_by, handled, notes)
                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
              (ticket.ticket_number, ticket.date_of_ticket, ticket.ticket_type, ticket.answered_by,
               ticket.response_time, ticket.claimed_by, ticket.closed_by, ticket.reviewed_by,
               ticket.handled, ticket.notes))
    conn.commit()
    conn.close()

def insert_interaction(interaction):
    """
    Inserts an interaction into the database and prints the interaction details for verification.
    """
    conn = sqlite3.connect('staff_management.db')
    c = conn.cursor()

    # Debug: Print interaction details before insertion
    print(f"Inserting interaction: Moderator: {interaction.moderator_name}, Date: {interaction.date_of_interaction}")

    c.execute('''INSERT INTO interactions (moderator_name, date_of_interaction, interaction_type)
                 VALUES (?, ?, ?)''',
              (interaction.moderator_name, interaction.date_of_interaction, interaction.interaction_type))
    conn.commit()
    conn.close()

# Function to export tickets, interactions, and staff data to Excel
def export_to_excel():
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return

    try:
        workbook = openpyxl.Workbook()

        # Tickets sheet
        tickets_sheet = workbook.active
        tickets_sheet.title = 'Tickets'
        tickets_sheet.append(['Ticket Number', 'Date', 'Type', 'Answered By', 'Claimed By', 'Response Time', 'Closed By', 'Reviewed By', 'Handled', 'Notes'])

        conn = sqlite3.connect('staff_management.db')
        c = conn.cursor()

        # Fetch and append ticket data
        c.execute('SELECT * FROM tickets')
        tickets = c.fetchall()
        for ticket in tickets:
            date_of_ticket = ticket[1]
            if date_of_ticket:
                try:
                    date_obj = datetime.strptime(date_of_ticket, "%Y-%m-%d")
                    date_of_ticket = date_obj.strftime("%m/%d/%Y")
                except ValueError:
                    date_of_ticket = date_of_ticket
            
            tickets_sheet.append([
                ticket[0],
                date_of_ticket,
                ticket[2],
                ticket[3],
                ticket[4],
                ticket[5],
                ticket[6],
                ticket[7],
                ticket[8],
                ticket[9]
            ])

        # Interactions sheet
        interactions_sheet = workbook.create_sheet('Interactions')
        interactions_sheet.append(['Moderator Name', 'Date of Interaction', 'Interaction Type'])

        c.execute('SELECT * FROM interactions')
        interactions = c.fetchall()
        for interaction in interactions:
            date_of_interaction = interaction[2]
            if date_of_interaction:
                try:
                    date_obj = datetime.strptime(date_of_interaction, "%Y-%m-%d")
                    date_of_interaction = date_obj.strftime("%m/%d/%Y")
                except ValueError:
                    date_of_interaction = date_of_interaction
            
            interactions_sheet.append([
                interaction[1],
                date_of_interaction,
                interaction[3]
            ])

        # Staff sheet
        staff_sheet = workbook.create_sheet('Staff')
        staff_sheet.append(['Name', 'Category'])

        c.execute('SELECT name, category FROM staff')
        staff = c.fetchall()
        for member in staff:
            staff_sheet.append(member)

        conn.close()

        workbook.save(file_path)
        messagebox.showinfo("Success", "Data successfully exported to Excel!")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to export data: {str(e)}")

# Trends and plotting functions
def get_ticket_trends(months):
    conn = sqlite3.connect('staff_management.db')
    c = conn.cursor()

    end_date = datetime.now()
    start_date = end_date - timedelta(days=months * 30)

    query = '''
        SELECT strftime('%m-%Y', date_of_ticket) AS month, COUNT(*), AVG(response_time)
        FROM tickets
        WHERE date_of_ticket BETWEEN ? AND ?
        GROUP BY month
        ORDER BY month
    '''
    c.execute(query, (start_date.strftime('%m-%d-%Y'), end_date.strftime('%m-%d-%Y')))
    ticket_data = c.fetchall()
    conn.close()

    return ticket_data

def get_interaction_trends(months, moderator=None):
    conn = sqlite3.connect('staff_management.db')
    c = conn.cursor()

    end_date = datetime.now()
    start_date = end_date - timedelta(days=months * 30)

    query = '''
        SELECT strftime('%m-%Y', date(date_of_interaction, 'start of month')) AS month, COUNT(*)
        FROM interactions
        WHERE date(date_of_interaction) BETWEEN ? AND ?
    '''
    params = [start_date.strftime('%m/%d/%Y'), end_date.strftime('%m/%d/%Y')]

    if moderator:
        query += ' AND moderator_name = ?'
        params.append(moderator)

    query += ' GROUP BY month ORDER BY month'
    c.execute(query, params)
    interaction_data = c.fetchall()
    conn.close()

    return interaction_data

def show_ticket_trends(months):
    ticket_data = get_ticket_trends(months)
    
    if not ticket_data:
        messagebox.showinfo("No Data", "No ticket data available for the selected period.")
        return

    months = [data[0] for data in ticket_data]
    ticket_counts = [data[1] for data in ticket_data]
    avg_response_times = [data[5] for data in ticket_data]

    fig, ax1 = plt.subplots()

    ax1.set_xlabel('Month')
    ax1.set_ylabel('Number of Tickets', color='tab:blue')
    ax1.plot(months, ticket_counts, color='tab:blue', label='Number of Tickets')
    ax1.tick_params(axis='y', labelcolor='tab:blue')

    ax2 = ax1.twinx()
    ax2.set_ylabel('Average Response Time (mins)', color='tab:red')
    ax2.plot(months, avg_response_times, color='tab:red', label='Average Response Time')
    ax2.tick_params(axis='y', labelcolor='tab:red')

    fig.tight_layout()
    plt.title(f'Ticket Trends for Last {months} Months')
    plt.show()

def show_interaction_trends(months, moderator=None):
    interaction_data = get_interaction_trends(months, moderator)
    
    if not interaction_data:
        messagebox.showinfo("No Data", "No interaction data available for the selected period.")
        return

    months = [data[0] for data in interaction_data]
    interaction_counts = [data[1] for data in interaction_data]

    plt.figure()
    plt.plot(months, interaction_counts, color='tab:green', label='Number of Interactions')
    plt.xlabel('Month')
    plt.ylabel('Number of Interactions')
    plt.title(f'Interaction Trends for Last {months} Months' + (f' (Moderator: {moderator})' if moderator else ''))
    plt.legend()
    plt.tight_layout()
    plt.show()

# GUI Application
class TicketApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Staff Management System")
        self.root.geometry("700x700")

        # Create or connect to the database
        create_database()

        # Create the main frame
        self.main_frame = tk.Frame(self.root)
        self.main_frame.pack(padx=10, pady=10)

        # Create buttons on the main page
        tk.Label(self.main_frame, text="Staff Management System", font=("Ariel", 16)).pack(pady=10)

        # Buttons for main options
        tk.Button(self.main_frame, text="Enter New Ticket", width=20, command=self.prompt_ticket_number).pack(pady=5)
        tk.Button(self.main_frame, text="Enter Interactions", width=20, command=self.enter_interactions).pack(pady=5)
        tk.Button(self.main_frame, text="Search Tickets", width=20, command=self.search_tickets).pack(pady=5)
        tk.Button(self.main_frame, text="Search Interactions", width=20, command=self.search_interactions).pack(pady=5)
        tk.Button(self.main_frame, text="View Ticket Trends (3 months)", width=30, command=lambda: self.view_ticket_trends(3)).pack(pady=5)
        tk.Button(self.main_frame, text="View Ticket Trends (6 months)", width=30, command=lambda: self.view_ticket_trends(6)).pack(pady=5)
        tk.Button(self.main_frame, text="View Interaction Trends (3 months)", width=30, command=lambda: self.view_interaction_trends(3)).pack(pady=5)
        tk.Button(self.main_frame, text="View Interaction Trends (6 months)", width=30, command=lambda: self.view_interaction_trends(6)).pack(pady=5)
        tk.Button(self.main_frame, text="Manage Allowed Names", width=20, command=self.manage_allowed_names).pack(pady=5)
        tk.Button(self.main_frame, text="Upload from Excel", width=20, command=upload_from_excel).pack(pady=5)
        tk.Button(self.main_frame, text="Export to Excel", width=20, command=export_to_excel).pack(pady=5)

    def prompt_ticket_number(self):
        ticket_window = tk.Toplevel(self.root)
        ticket_window.title("Enter New Ticket - Ticket Number")

        tk.Label(ticket_window, text="Ticket Number").grid(row=0, column=0)
        ticket_number_entry = tk.Entry(ticket_window)
        ticket_number_entry.grid(row=0, column=1)

        def check_ticket_number():
            ticket_num = ticket_number_entry.get()

            if ticket_number_exists(ticket_num):
                messagebox.showerror("Error", f"Ticket number {ticket_num} already exists. Please enter a different ticket number.")
                ticket_number_entry.delete(0, tk.END)
            else:
                ticket_window.destroy()
                self.enter_ticket(ticket_num)  # Proceed to enter the rest of the ticket details

        tk.Button(ticket_window, text="Next", command=check_ticket_number).grid(row=1, column=0, columnspan=2, pady=10)

    def enter_ticket(self, ticket_num):
        ticket_window = tk.Toplevel(self.root)
        ticket_window.title("Enter New Ticket")

        tk.Label(ticket_window, text="Ticket Number").grid(row=0, column=0)
        ticket_number = tk.Entry(ticket_window)
        ticket_number.insert(0, ticket_num)
        ticket_number.config(state="readonly")
        ticket_number.grid(row=0, column=1)

        tk.Label(ticket_window, text="Date (MM/DD/YYYY)").grid(row=1, column=0)
        ticket_date = tk.Entry(ticket_window)
        ticket_date.grid(row=1, column=1)

        tk.Label(ticket_window, text="Ticket Type").grid(row=2, column=0)
        ticket_type = tk.Entry(ticket_window)
        ticket_type.grid(row=2, column=1)

        staff = [name for name, _ in get_all_staff()]

        tk.Label(ticket_window, text="Answered By").grid(row=3, column=0)
        answered_by = ttk.Combobox(ticket_window, values=staff)
        answered_by.grid(row=3, column=1)

        tk.Label(ticket_window, text="Response Time (mins)").grid(row=4, column=0)
        response_time = tk.Entry(ticket_window)
        response_time.grid(row=4, column=1)

        tk.Label(ticket_window, text="Claimed By").grid(row=5, column=0)
        claimed_by = ttk.Combobox(ticket_window, values=staff)
        claimed_by.grid(row=5, column=1)

        tk.Label(ticket_window, text="Closed By").grid(row=6, column=0)
        closed_by = ttk.Combobox(ticket_window, values=staff)
        closed_by.grid(row=6, column=1)

        tk.Label(ticket_window, text="Reviewed By").grid(row=7, column=0)
        reviewed_by = tk.Entry(ticket_window)
        reviewed_by.grid(row=7, column=1)

        tk.Label(ticket_window, text="Handled").grid(row=8, column=0)
        handled = tk.Entry(ticket_window)
        handled.grid(row=8, column=1)

        tk.Label(ticket_window, text="Notes").grid(row=9, column=0)
        notes = tk.Entry(ticket_window)
        notes.grid(row=9, column=1)

        def save_ticket():
            ticket = Ticket(
                ticket_number.get(),
                ticket_date.get(),
                ticket_type.get(),
                answered_by.get(),
                response_time.get(),
                claimed_by.get(),
                closed_by.get(),
                reviewed_by.get(),
                handled.get(),
                notes.get()
            )
            insert_ticket(ticket)
            messagebox.showinfo("Success", "Ticket saved successfully!")
            ticket_window.destroy()

        tk.Button(ticket_window, text="Save Ticket", command=save_ticket).grid(row=10, column=0, columnspan=2, pady=10)

    def enter_interactions(self):
        interaction_window = tk.Toplevel(self.root)
        interaction_window.title("Enter New Interaction")

        staff = [name for name, category in get_all_staff() if category in CATEGORY_RANKS]

        tk.Label(interaction_window, text="Moderator Name").grid(row=0, column=0)
        moderator_name = ttk.Combobox(interaction_window, values=staff)
        moderator_name.grid(row=0, column=1)

        tk.Label(interaction_window, text="Date (MM/DD/YYYY)").grid(row=1, column=0)
        interaction_date = tk.Entry(interaction_window)
        interaction_date.grid(row=1, column=1)

        tk.Label(interaction_window, text="Interaction Type").grid(row=2, column=0)
        interaction_type = tk.Entry(interaction_window)
        interaction_type.grid(row=2, column=1)

        def save_interaction():
            interaction = ModeratorInteraction(
                moderator_name.get(),
                interaction_date.get(),
                interaction_type.get()
            )
            insert_interaction(interaction)
            messagebox.showinfo("Success", "Interaction saved successfully!")
            interaction_window.destroy()

        tk.Button(interaction_window, text="Save Interaction", command=save_interaction).grid(row=3, column=0, columnspan=2, pady=10)

    def search_tickets(self):
        search_window = tk.Toplevel(self.root)
        search_window.title("Search Tickets")

        staff = [name for name, _ in get_all_staff()]

        tk.Label(search_window, text="Moderator Name").grid(row=0, column=0)
        self.moderator_name = ttk.Combobox(search_window, values=staff)
        self.moderator_name.grid(row=0, column=1)

        tk.Label(search_window, text="Select Month").grid(row=1, column=0)
        months = [str(i).zfill(2) for i in range(1, 13)]
        self.month_combobox = ttk.Combobox(search_window, values=months, state="readonly")
        self.month_combobox.grid(row=1, column=1)

        tk.Label(search_window, text="Select Year").grid(row=2, column=0)
        years = [str(year) for year in range(2023, datetime.now().year + 1)]
        self.year_combobox = ttk.Combobox(search_window, values=years, state="readonly")
        self.year_combobox.grid(row=2, column=1)

        tk.Button(search_window, text="Search", command=self.perform_ticket_search).grid(row=3, column=0, columnspan=2, pady=10)

    def perform_ticket_search(self):
        moderator = self.moderator_name.get().strip()
        selected_month = self.month_combobox.get().strip()
        selected_year = self.year_combobox.get().strip()

        tickets = get_filtered_tickets(moderator_name=moderator, selected_month=selected_month, selected_year=selected_year)

        if not tickets:
            messagebox.showinfo("No Results", "No tickets found for the given criteria.")
            return

        result_window = tk.Toplevel(self.root)
        result_window.title(f"Search Results - Total: {len(tickets)} Tickets")

        tree = ttk.Treeview(result_window, columns=('Ticket Number', 'Date', 'Type', 'Answered By', 'Claimed By', 'Response Time', 'Notes'), show='headings')
        tree.heading('Ticket Number', text='Ticket Number')
        tree.heading('Date', text='Date')
        tree.heading('Type', text='Type')
        tree.heading('Answered By', text='Answered By')
        tree.heading('Claimed By', text='Claimed By')
        tree.heading('Response Time', text='Response Time')
        tree.heading('Notes', text='Notes')
        tree.pack(fill=tk.BOTH, expand=True)

        for ticket in tickets:
            date_str = ticket[1]
            try:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                month_name = date_obj.strftime("%B")
                formatted_date = f"{month_name} {date_obj.day}, {date_obj.year}"
            except ValueError:
                formatted_date = date_str

            tree.insert('', 'end', values=(ticket[0], formatted_date, ticket[2], ticket[3], ticket[5], ticket[4], ticket[9]))

        average_response_time = calculate_average_response_time(tickets)
        tk.Label(result_window, text=f"Average Response Time: {average_response_time:.2f} mins").pack()

    def search_interactions(self):
        search_window = tk.Toplevel(self.root)
        search_window.title("Search Interactions")

        staff = [name for name, _ in get_all_staff()]

        tk.Label(search_window, text="Moderator Name").grid(row=0, column=0)
        self.moderator_name = ttk.Combobox(search_window, values=staff)
        self.moderator_name.grid(row=0, column=1)

        tk.Label(search_window, text="Select Month").grid(row=1, column=0)
        months = [str(i).zfill(2) for i in range(1, 13)]
        self.month_combobox = ttk.Combobox(search_window, values=months, state="readonly")
        self.month_combobox.grid(row=1, column=1)

        tk.Label(search_window, text="Select Year").grid(row=2, column=0)
        years = [str(year) for year in range(2023, datetime.now().year + 1)]
        self.year_combobox = ttk.Combobox(search_window, values=years, state="readonly")
        self.year_combobox.grid(row=2, column=1)

        tk.Button(search_window, text="Search", command=self.perform_interaction_search).grid(row=3, column=0, columnspan=2, pady=10)

    def perform_interaction_search(self):
        moderator = self.moderator_name.get().strip()
        selected_month = self.month_combobox.get().strip()
        selected_year = self.year_combobox.get().strip()

        interactions = get_filtered_interactions(moderator, selected_month, selected_year)

        if not interactions:
            messagebox.showinfo("No Results", "No interactions found for the given criteria.")
            return

        result_window = tk.Toplevel(self.root)
        result_window.title(f"Search Results - Total: {len(interactions)} Interactions")

        tree = ttk.Treeview(result_window, columns=('Moderator', 'Date', 'Type'), show='headings')
        tree.heading('Moderator', text='Moderator')
        tree.heading('Date', text='Date')
        tree.heading('Type', text='Type')
        tree.pack(fill=tk.BOTH, expand=True)

        for interaction in interactions:
            tree.insert('', 'end', values=(interaction[1], interaction[2], interaction[3]))

        tk.Label(result_window, text=f"Total Interactions: {len(interactions)}").pack()

        self.show_interaction_trend_graph(moderator)

    def show_interaction_trend_graph(self, moderator):
        interaction_data = get_interaction_trends(6, moderator)

        if not interaction_data:
            messagebox.showinfo("No Trend Data", f"No interaction trend data available for {moderator}.")
            return

        months = [data[0] for data in interaction_data]
        interaction_counts = [data[1] for data in interaction_data]

        plt.figure()
        plt.plot(months, interaction_counts, color='tab:green', marker='o', label=f'Interactions for {moderator}')
        plt.xlabel('Month')
        plt.ylabel('Number of Interactions')
        plt.title(f'Interaction Trends (Last 6 Months) for {moderator}')
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.legend()
        plt.show()

    def view_ticket_trends(self, months):
        show_ticket_trends(months)

    def view_interaction_trends(self, months):
        show_interaction_trends(months)

    def manage_allowed_names(self):
        name_window = tk.Toplevel(self.root)
        name_window.title("Manage Allowed Names")

        tk.Label(name_window, text="Allowed Names").grid(row=0, column=0)
        allowed_names_list = tk.Listbox(name_window, height=10)
        allowed_names_list.grid(row=1, column=0, columnspan=2, pady=5)

        staff = get_all_staff()
        for name, category in staff:
            allowed_names_list.insert(tk.END, f"{name} - {category}")

        tk.Label(name_window, text="Add New Name:").grid(row=2, column=0)
        name_entry = tk.Entry(name_window)
        name_entry.grid(row=2, column=1)

        tk.Label(name_window, text="Category:").grid(row=3, column=0)
        category_combobox = ttk.Combobox(name_window, values=CATEGORY_RANKS)
        category_combobox.grid(row=3, column=1)

        def add_name():
            new_name = name_entry.get().strip()
            selected_category = category_combobox.get().strip()
            if new_name and selected_category:
                conn = sqlite3.connect('staff_management.db')
                c = conn.cursor()
                c.execute("INSERT INTO staff (name, category) VALUES (?, ?)", (new_name, selected_category))
                conn.commit()
                conn.close()
                allowed_names_list.insert(tk.END, f"{new_name} - {selected_category}")
                name_entry.delete(0, tk.END)

        def remove_name():
            selected_name = allowed_names_list.get(tk.ACTIVE)
            if selected_name:
                name, category = selected_name.split(" - ")
                confirm = messagebox.askyesno("Confirm Removal", f"Are you sure you want to remove '{name}' from '{category}'?")
                if confirm:
                    conn = sqlite3.connect('staff_management.db')
                    c = conn.cursor()
                    c.execute("DELETE FROM staff WHERE name = ?", (name,))
                    conn.commit()
                    conn.close()
                    allowed_names_list.delete(tk.ACTIVE)
                    messagebox.showinfo("Success", f"{name} has been removed from {category}.")

        tk.Button(name_window, text="Add", command=add_name).grid(row=4, column=0)
        tk.Button(name_window, text="Remove", command=remove_name).grid(row=4, column=1)

if __name__ == "__main__":
    root = tk.Tk()
    app = TicketApp(root)
    root.mainloop()
